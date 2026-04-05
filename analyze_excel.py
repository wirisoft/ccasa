# -*- coding: utf-8 -*-
"""Analiza archivos Excel del proyecto y extrae hojas, columnas y muestras."""
import os
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl")
    exit(1)

from excel_file_manifest import BASE, FILES

def cell_value(c):
    if c is None:
        return ""
    v = c.value
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return str(v)
    return str(v).strip() if isinstance(v, str) else v

def analyze_sheet(ws, max_header_cols=30, max_sample_rows=5):
    rows = list(ws.iter_rows(max_row=ws.max_row or 1, max_col=ws.max_column or 1))
    if not rows:
        return {"headers": [], "sample_rows": [], "max_row": 0, "max_col": 0}
    headers = [cell_value(rows[0][j]) for j in range(min(len(rows[0]), max_header_cols))]
    sample = []
    for i in range(1, min(len(rows), 1 + max_sample_rows)):
        sample.append([cell_value(rows[i][j]) for j in range(min(len(rows[i]), max_header_cols))])
    return {
        "headers": headers,
        "sample_rows": sample,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
    }

def analyze_file(relpath):
    path = BASE / relpath
    if not path.exists():
        return {"error": "File not found", "path": str(path)}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = analyze_sheet(ws, max_header_cols=40, max_sample_rows=8)
        wb.close()
        return {"sheets": sheets, "path": str(path)}
    except Exception as e:
        return {"error": str(e), "path": str(path)}

def main():
    results = {}
    for rel in FILES:
        key = rel.replace("/", "_").replace(" ", "_")[:80]
        results[key] = analyze_file(rel)
    out = Path(__file__).resolve().parent / "excel_analysis.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print("Written:", out)

if __name__ == "__main__":
    main()
