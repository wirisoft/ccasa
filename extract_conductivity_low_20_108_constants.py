# -*- coding: utf-8 -*-
"""
Lee celdas de constantes del libro 20-108-01 --CONDUCTIVIDAD BAJAS.xlsx (hoja YYYYMMDD)
y escribe conductivity_low_20_108_constants.json en la raíz del repo.

Misma rejilla de referencia que extract_conductivity_high_20_108_constants.py.

Uso:
  python extract_conductivity_low_20_108_constants.py [ruta_al_xlsx]

Requiere: pip install openpyxl
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

OUT = Path(__file__).resolve().parent / "conductivity_low_20_108_constants.json"

CELLS_BLOCK1 = [
    "B10",
    "B24",
    "C24",
    "C25",
    "F24",
    "D28",
    "F28",
]
CELLS_BLOCK2 = [
    "B40",
    "B54",
    "C54",
    "C55",
    "F54",
    "D58",
    "F58",
]


def main():
    default = (
        Path(__file__).resolve().parent
        / "filesproyect"
        / "3-CONDUCTIVIDAD BAJA"
        / "20-108-01 --CONDUCTIVIDAD BAJAS.xlsx"
    )
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not path.is_file():
        print(f"No existe el archivo: {path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if len(name.strip()) == 8 and name.strip().isdigit():
            sheet_name = name.strip()
            break
    if sheet_name is None:
        print("No se encontró hoja YYYYMMDD", file=sys.stderr)
        sys.exit(3)

    ws = wb[sheet_name]

    def cell_value(addr):
        c = ws[addr]
        v = c.value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return str(v)
        return str(v).strip()

    block1 = {a: cell_value(a) for a in CELLS_BLOCK1}
    block2 = {a: cell_value(a) for a in CELLS_BLOCK2}

    payload = {
        "source_file": str(path),
        "sheet": sheet_name,
        "block1": block1,
        "block2": block2,
        "constants": {
            "C25": block1.get("C25"),
            "B24": block1.get("B24"),
            "C24": block1.get("C24"),
            "F24": block1.get("F24"),
            "D28": block1.get("D28"),
            "F28": block1.get("F28"),
        },
    }

    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Escrito: {OUT}")


if __name__ == "__main__":
    main()
