# -*- coding: utf-8 -*-
"""
Extrae celdas con fórmulas de los Excel del proyecto (data_only=False).
Salida: excel_formulas.json en la raíz del repo.

Requiere: pip install openpyxl
"""
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from excel_file_manifest import BASE, FILES

OUT = Path(__file__).resolve().parent / "excel_formulas.json"


def is_formula_cell(cell):
    v = cell.value
    if v is None:
        return False
    if isinstance(v, str) and v.startswith("="):
        return True
    return getattr(cell, "data_type", None) == "f"


def collect_formulas(ws, max_formulas_per_sheet=50000):
    """Itera celdas con fórmula; límite por hoja (subir si alguna hoja supera el tope y regenerar excel_formulas.json)."""
    found = []
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if not is_formula_cell(cell):
                continue
            if count >= max_formulas_per_sheet:
                return found, True
            coord = cell.coordinate
            val = cell.value
            found.append({"cell": coord, "formula": val if isinstance(val, str) else str(val)})
            count += 1
    return found, False


def analyze_workbook(path: Path):
    # read_only=True permite iterar hojas grandes; data_only=False conserva fórmulas
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets_out = {}
    truncated = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            formulas, cut = collect_formulas(ws)
            sheets_out[name] = {"formulas": formulas, "count": len(formulas)}
            if cut:
                truncated[name] = True
    finally:
        wb.close()
    return {"sheets": sheets_out, "truncated": truncated, "path": str(path)}


def main():
    root = Path(__file__).resolve().parent
    base = Path(os.environ.get("CCASA_FILES_ROOT", str(BASE)))
    if not base.is_absolute():
        base = root / base
    results = {}
    for rel in FILES:
        path = base / rel
        key = rel.replace("/", "_").replace(" ", "_")[:80]
        if not path.exists():
            results[key] = {"error": "File not found", "path": str(path)}
            continue
        try:
            results[key] = analyze_workbook(path)
        except Exception as e:
            results[key] = {"error": str(e), "path": str(path)}
    OUT.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Written:", OUT)


if __name__ == "__main__":
    main()
